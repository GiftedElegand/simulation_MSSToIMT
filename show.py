from math import pi

import numpy as np
from matplotlib import pyplot as plt
from openpyxl import Workbook

from IMT_Simulation import gain_calc
import statsmodels.api as sm


def showS_h():
    y = []
    list=[]
    for i in range(-180, 180):
        a = gain_calc(pi / 2, 0, pi / 2, i / 180 * pi)
        if a < -200:
            y.append(y[len(y) - 1])
        else:
            y.append(a)
        list.append([90,0,90,i,a])
    x = range(-180, 180)
    # y1 = [0.86, 0.85, 0.853, 0.849, 0.83]
    # plt.plot(x, y, 'ro-')
    # plt.plot(x, y1, 'bo-')
    plt.figure(figsize=(12, 8), dpi=80)
    plt.xlim(-180, 180)  # 限定横轴的范围
    plt.ylim(min(y) - 10, max(y) + 10)  # 限定纵轴的范围
    plt.xticks(x[::20])
    # plt.yticks(y[::10])
    # 添加网格信息
    plt.grid(True, linestyle='--', alpha=0.5)
    plt.plot(x, y, marker='.', mec='r', mfc='w', label=u'UE_v=90, UE_h=0, S_v=90, S_h=x')
    plt.legend()  # 让图例生效
    # plt.xticks(x, names, rotation=45)
    plt.margins(0)
    plt.subplots_adjust(bottom=0.1)
    plt.xlabel(u"HIBS h")  # X轴标签
    plt.ylabel("A_A")  # Y轴标签
    plt.title("A simple plot")  # 标题

    plt.show()
    return list

def showS_v():
    y = []
    list = []
    for i in range(0, 180):
        a = gain_calc(0, 0, i / 180 * pi, 0)
        y.append(a)
        list.append([0, 0, i, 0, a])
    x = range(0, 180)
    plt.figure(figsize=(12, 8), dpi=80)
    plt.xlim(0, 180)  # 限定横轴的范围
    plt.ylim(min(y) - 10, max(y) + 10)  # 限定纵轴的范围
    plt.xticks(x[::20])
    # 添加网格信息
    plt.grid(True, linestyle='--', alpha=0.5)
    plt.plot(x, y, marker='.', mec='r', mfc='w', label=u'UE_v=0, UE_h=0, S_v=x, S_h=0')
    plt.legend()  # 让图例生效
    plt.margins(0)
    plt.subplots_adjust(bottom=0.1)
    plt.xlabel(u"HIBS v")  # X轴标签
    plt.ylabel("A_A")  # Y轴标签
    plt.title("A simple plot")  # 标题

    plt.show()
    return list



# 创建Excel文件
def create_excel(data_list,filename):
    # 注：以下写入excel表是自动新建excel表，然后写入。如果excel表已经存在，再执行下面语句，会将原本excel表内的数据清空，然后将内容写入。
    wb = Workbook()  # 创建工作薄对象
    ws = wb['Sheet']  # 创建子表 （注意：Sheet中的S需要大写，不然可能会出错）
    ws.cell(1, 1, "UE的倾角")
    ws.cell(1, 2, "UE方位角")
    ws.cell(1, 3, "HIBS倾角")
    ws.cell(1, 4, "HIBS方位角")
    ws.cell(1, 5, "增益")
    ws.cell(1, 6, "PFD")
    for i in range(len(data_list)):
        ws.cell(row=i + 2, column=1).value = data_list[i][0]
        ws.cell(row=i + 2, column=2).value = data_list[i][1]
        ws.cell(row=i + 2, column=3).value = data_list[i][2]
        ws.cell(row=i + 2, column=4).value = data_list[i][3]
        ws.cell(row=i + 2, column=5).value = data_list[i][4]  # 将数据data写入excel中的第i行第j列
        ws.cell(row=i + 2, column=6).value = data_list[i][5]
    wb.save(filename+'.xlsx')  # 保存excel，保存的文件名是MY_EXCEL.xlsx


def CDF_plt(data,pltname,SeparationDistance):
    ecdf = sm.distributions.ECDF(data)
    # 等差数列，用于绘制X轴数据
    x = np.linspace(min(data), max(data))
    # x轴数据上值对应的累计密度概率
    y = ecdf(x)
    # 绘制阶梯图
    plt.ylim(0, 1.05)
    plt.title(f"{pltname} Interference,Separation Distance={SeparationDistance/1000}km")
    plt.plot(x, y)
    # IMT系统为-6，FS为-126
    mark_x=-126
    mark_y=ecdf(mark_x)
    plt.plot(mark_x, mark_y, color="blue", marker="^")
    plt.plot([mark_x, mark_x], [0, mark_y], ls='--', color='gray', lw=1)
    plt.plot(mark_x, 0, '|', color='k', ms=10)
    plt.xlabel(f"CDF P(X ≤ {mark_x})={mark_y}")
    # plt.xlabel(f"CDF P(X ≤ -126)={ecdf(-126)}")
    # idx = np.argmin(np.abs(y - 0.8))
    # plt.xlabel(f"CDF P(Y ≤ 0.8)={x[idx]}")
    # plt.title(pltname)
    plt.savefig(pltname+".jpg")
    plt.show()

